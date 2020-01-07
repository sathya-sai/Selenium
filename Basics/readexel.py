"""
    :summary This is python 3.7 supported selenium 3.141.0

    :since January 2020

    :author Sathya Sai M

    :keyword Python, selenium to read from EXEL file
"""
import openpyxl


class Exel:
    def exel(self):
        path = "/home/admin1/Downloads/Financial Sample.xlsx"

        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        rows = sheet.max_row
        cols = sheet.max_column

        print(rows)
        print(cols)

        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                print(sheet.cell(row=r, column=c).value, end="     ")

            print()


if __name__ == '__main__':
    e = Exel()
    e.exel()
